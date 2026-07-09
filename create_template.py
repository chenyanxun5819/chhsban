import openpyxl
from pathlib import Path

# 创建 template.xlsx
output_path = r"D:\chhsban\teacher-management\examples\template.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Teachers"

# 添加列标题
headers = ["department", "School ID", "Name", "email"]
for col_idx, header in enumerate(headers, 1):
    ws.cell(1, col_idx, header)

# 添加示例数据
sample_data = [
    ["华文 Chinese", "T119", "谭长咏", "ecchhs014@chhsban.edu.my"],
    ["数学 Maths", "T001", "李明", "ecchhs001@chhsban.edu.my"],
]
for row_idx, row_data in enumerate(sample_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row_idx, col_idx, value)

# 设置列宽
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 30

wb.save(output_path)
print(f"✅ 创建 template.xlsx: {output_path}")
