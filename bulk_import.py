import openpyxl
import requests
import json

# 读取 Excel 文件
xlsx_file = r"D:\chhsban\chhsban-acadoc\20240802 全体教师名单 电话号码 教育账号.xlsx"
wb = openpyxl.load_workbook(xlsx_file)
ws = wb.active

# 提取数据
teachers = []
for row_idx in range(2, ws.max_row + 1):
    department = ws.cell(row_idx, 1).value
    school_id = ws.cell(row_idx, 2).value
    name = ws.cell(row_idx, 3).value
    email = ws.cell(row_idx, 4).value
    
    if school_id and name and email:  # 验证必填字段
        teachers.append({
            "teacher_id": school_id,
            "name_cn": name,
            "department": department or "",
            "email": email
        })

print(f"📊 准备导入 {len(teachers)} 位教师\n")

# 调用批量导入 API
api_url = "https://teacher-management.astcws.workers.dev/api/teachers/import"
headers = {
    "Content-Type": "application/json",
    "Authorization": "Bearer test_key"
}

try:
    print("🚀 开始导入...")
    response = requests.post(api_url, json={"teachers": teachers}, headers=headers)
    
    if response.status_code == 200:
        data = response.json()
        results = data.get("data", {})
        print(f"\n✅ 导入完成!")
        print(f"  • 新增: {results.get('created', 0)} 位")
        print(f"  • 更新: {results.get('updated', 0)} 位")
        print(f"  • 跳过: {results.get('skipped', 0)} 位")
        print(f"  • 错误: {results.get('errors', [])}")
    else:
        print(f"❌ 导入失败: {response.status_code}")
        print(response.text)
except Exception as error:
    print(f"❌ 错误: {error}")
