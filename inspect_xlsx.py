#!/usr/bin/env python3
import openpyxl
from pathlib import Path

# 找到上传的文件
xlsx_file = r"D:\chhsban\chhsban-acadoc\20240802 全体教师名单 电话号码 教育账号.xlsx"

if not Path(xlsx_file).exists():
    print(f"❌ 找不到文件: {xlsx_file}")
    exit(1)

# 读取 Excel 文件
wb = openpyxl.load_workbook(xlsx_file)
ws = wb.active

print(f"📊 文件: {Path(xlsx_file).name}")
print(f"📏 行数: {ws.max_row}, 列数: {ws.max_column}\n")

# 显示列标题
print("📋 列标题:")
headers = []
for col_idx in range(1, ws.max_column + 1):
    cell = ws.cell(1, col_idx)
    headers.append(cell.value)
    print(f"  {chr(64+col_idx)}: {cell.value}")

# 显示前 10 行数据
print("\n📌 前 10 行数据:")
for row_idx in range(2, min(12, ws.max_row + 1)):
    row_data = []
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row_idx, col_idx)
        row_data.append(str(cell.value)[:20] if cell.value else '')
    print(f"  {row_idx}: {' | '.join(row_data)}")

# 统计有效数据行
valid_count = 0
for row_idx in range(2, ws.max_row + 1):
    if ws.cell(row_idx, 1).value:  # 检查第一列
        valid_count += 1

print(f"\n✅ 总有效数据行: {valid_count}")
