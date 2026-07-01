"""
将 Excel 转换为 JSON 并上传到 Cloudflare KV
一次性操作，为 Worker 准备数据
"""

import json
from pathlib import Path
import requests
from openpyxl import load_workbook
from dotenv import load_dotenv
import os

# 加载环境变量
load_dotenv()

CLOUDFLARE_ACCOUNT_ID = os.getenv('CLOUDFLARE_ACCOUNT_ID')
CLOUDFLARE_NAMESPACE_ID = os.getenv('CLOUDFLARE_NAMESPACE_ID')
CLOUDFLARE_API_TOKEN = os.getenv('CLOUDFLARE_API_TOKEN')

if not all([CLOUDFLARE_ACCOUNT_ID, CLOUDFLARE_NAMESPACE_ID, CLOUDFLARE_API_TOKEN]):
    raise ValueError("Missing Cloudflare credentials in .env")

BASE_URL = f"https://api.cloudflare.com/client/v4/accounts/{CLOUDFLARE_ACCOUNT_ID}/storage/kv/namespaces/{CLOUDFLARE_NAMESPACE_ID}"

HEADERS = {
    "Authorization": f"Bearer {CLOUDFLARE_API_TOKEN}",
    "Content-Type": "application/json"
}


def excel_to_json():
    """将 Excel 文件转换为 JSON 字典"""
    print("📖 读取 Excel 文件...")
    
    script_dir = Path(__file__).parent
    excel_file = script_dir / "2026_hostelList.xlsx"
    
    if not excel_file.exists():
        raise FileNotFoundError(f"找不到 Excel 文件: {excel_file}")
    
    print(f"   文件路径: {excel_file}")
    
    # 打开 Excel
    wb = load_workbook(excel_file)
    
    # 查找工作表（通常是时间戳名称）
    sheet_names = wb.sheetnames
    print(f"   找到工作表: {sheet_names}")
    
    # 使用第一个工作表
    ws = wb[sheet_names[0]]
    print(f"   使用工作表: {sheet_names[0]}")
    
    # 提取数据
    # 列 C (索引 2) = studentID
    # 列 F (索引 5) = Gender/Boarding
    excel_map = {}
    
    print("   提取数据...")
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            student_id = str(row[2]).strip() if row[2] else None  # 列 C
            gender_boarding = str(row[5]).strip() if row[5] else None  # 列 F
            
            if student_id and gender_boarding and student_id != 'None':
                excel_map[student_id] = gender_boarding
        except Exception as e:
            print(f"   行 {row_idx} 出错: {e}")
            continue
    
    print(f"✅ 提取完成: {len(excel_map)} 条记录")
    
    # 显示样本数据
    print("\n   样本数据:")
    for idx, (sid, gb) in enumerate(list(excel_map.items())[:5]):
        print(f"      {sid} → {gb}")
    
    return excel_map


def upload_to_kv(data):
    """上传到 Cloudflare KV"""
    print("\n☁️ 上传到 Cloudflare KV...")
    
    key = "excel_gender_boarding_map"
    
    try:
        url = f"{BASE_URL}/values/{key}"
        
        # 准备数据
        json_data = json.dumps(data, ensure_ascii=False)
        
        print(f"   键名: {key}")
        print(f"   数据大小: {len(json_data)} 字节")
        
        # 上传
        response = requests.put(
            url,
            headers=HEADERS,
            data=json_data.encode('utf-8'),
            timeout=30
        )
        
        if response.status_code == 200:
            print(f"✅ 上传成功！")
            print(f"   记录数: {len(data)}")
            print(f"   大小: {len(json_data) / 1024:.2f} KB")
            return True
        else:
            print(f"❌ 上传失败: {response.status_code}")
            print(f"   响应: {response.text[:200]}")
            return False
            
    except Exception as e:
        print(f"❌ 上传错误: {str(e)}")
        return False


def verify_upload():
    """验证上传是否成功"""
    print("\n🔍 验证上传...")
    
    try:
        url = f"{BASE_URL}/values/excel_gender_boarding_map"
        
        response = requests.get(url, headers=HEADERS, timeout=10)
        
        if response.status_code == 200:
            data = response.json()
            count = len(data) if isinstance(data, dict) else 1
            print(f"✅ 验证成功！KV 中有 {count} 条记录")
            
            # 显示样本
            if isinstance(data, dict):
                print("   样本数据:")
                for idx, (sid, gb) in enumerate(list(data.items())[:3]):
                    print(f"      {sid} → {gb}")
            
            return True
        else:
            print(f"❌ 验证失败: {response.status_code}")
            return False
            
    except Exception as e:
        print(f"❌ 验证错误: {str(e)}")
        return False


def main():
    print("="*60)
    print("🚀 Excel → KV 上传流程")
    print("="*60 + "\n")
    
    try:
        # 步骤 1: 读取 Excel
        excel_data = excel_to_json()
        
        # 步骤 2: 上传到 KV
        if not upload_to_kv(excel_data):
            print("\n❌ 上传失败，请检查凭证")
            return False
        
        # 步骤 3: 验证
        if not verify_upload():
            print("\n⚠️ 无法验证上传，但可能已成功")
        
        print("\n" + "="*60)
        print("✅ 完成！Excel 数据已准备好用于 Worker")
        print("="*60)
        
        return True
        
    except Exception as e:
        print(f"\n❌ 错误: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    import sys
    success = main()
    sys.exit(0 if success else 1)
