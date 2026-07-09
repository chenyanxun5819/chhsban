import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# 创建真正的 Excel 文件
output_path = r"D:\chhsban\teacher-management\examples\template.xlsx"

# 创建工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Teachers"

# 添加列标题（粗体和背景色）
headers = ["department", "School ID", "Name", "email"]
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(1, col_idx, header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# 添加示例数据
sample_data = [
    ["华文 Chinese", "T119", "谭长咏", "ecchhs014@chhsban.edu.my"],
    ["数学 Maths", "T001", "李明", "ecchhs001@chhsban.edu.my"],
    ["英文 English", "T100", "王美玲", "ecchhs100@chhsban.edu.my"],
]

for row_idx, row_data in enumerate(sample_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row_idx, col_idx, value)

# 设置列宽
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 30

# 设置行高
ws.row_dimensions[1].height = 25

# 保存文件
wb.save(output_path)
print(f"✅ 创建真正的 Excel 文件: {output_path}")
